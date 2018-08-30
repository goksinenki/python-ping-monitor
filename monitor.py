#!/usr/bin/env python
# -*- coding: utf-8 -*-

import socket
import sys
import cmd
import time
import re
import os
from os import system
import subprocess
import commands
import smtplib
import string
import xlsxwriter
import csv
import openpyxl

from email import Encoders
from email.MIMEBase import MIMEBase
from email.MIMEMultipart import MIMEMultipart
from email.Utils import formatdate
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell


buff = ''
resp = ''

my_file = open("ip.txt", "rb")
ths = open('pingsonuclari.csv', 'w')
s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)  #Create a TCP/IP socket

for line in my_file:
        l = [i.strip() for i in line.split(' ')]
        Hostname = l[0]
        IP = l[1]
        try:
            output = subprocess.check_output('ping -n 1 %s'%(IP),shell=True)
        except Exception:
            output = None
       #print output		
        if 'TTL' in output:
			print '%s %s is UP'%(Hostname,IP)
			ths.write('%s %s is UP\n'%(Hostname,IP))
        else:
			print '%s %s is DOWN'%(Hostname,IP)
			ths.write('%s %s is DOWN\n'%(Hostname,IP))
	
	
ths.close()
my_file.close()



reload(sys)
sys.setdefaultencoding('utf8')

redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')

greenFill = PatternFill(start_color='FF00FF00',
                   end_color='FF00FF00',
                   fill_type='solid')


if __name__ == '__main__':
    workbook = Workbook()
    worksheet = workbook.active
    with open('pingsonuclari.csv', 'r') as f:
        reader = csv.reader(f)
        worksheet.column_dimensions["A"].width = 60
        for r, row in enumerate(reader):
            for c, col in enumerate(row):
                for idx, val in enumerate(col.split('\\s+')):
                    cell = worksheet.cell(row=r+1, column=c+1)
                    cell.value = val
                    if 'DOWN' in val:
							cell.fill = redFill
                    else :
						cell.fill = greenFill
    workbook.save('network_cihazlari_ping_raporu.xlsx')



filePath = r'network_cihazlari_ping_raporu.xlsx'
 
def sendEmail(TO = "admin1@company.com",
              CC = "admin2@company.com",
              FROM="pingmonitor@company.com"):
    HOST = "mail.company.com"
 
    msg = MIMEMultipart()
    msg["From"] = FROM
    msg["To"] = TO
    msg["Cc"] = CC
    msg["Subject"] = "Üretim öncesi network switch ping raporu!".decode("utf-8")
    msg['Date']    = formatdate(localtime=True)
 
    # attach a file
    part = MIMEBase('application', "octet-stream")
    part.set_payload( open(filePath,"rb").read() )
    Encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment; filename="%s"' % os.path.basename(filePath))
    msg.attach(part)
 
    server = smtplib.SMTP(HOST)
    # server.login(username, password)  # optional
 
    try:
        failed = server.sendmail(FROM, TO, msg.as_string())
        server.close()
    except Exception, e:
        errorMsg = "Unable to send email. Error: %s" % str(e)
 
if __name__ == "__main__":
    sendEmail()
	
k=input("press close to exit") 
